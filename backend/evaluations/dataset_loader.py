from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


EXPECTED = {"Happy Path": 20, "Edge Cases": 12, "Known Failures": 6, "Adversarial": 2}
CANONICAL_CATEGORIES = {name.casefold(): name for name in EXPECTED}


def load_golden_dataset(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Golden Dataset"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value).strip() for value in rows[0]]
    return [dict(zip(headers, row, strict=False)) for row in rows[1:] if any(value is not None for value in row)]


def validate_distribution(rows: list[dict[str, object]]) -> Counter:
    if not rows:
        raise ValueError("Golden Dataset is empty")
    category_key = next((key for key in rows[0] if "category" in key.lower() or "scenario type" in key.lower()), None)
    if category_key is None:
        raise ValueError("Golden Dataset has no category column")
    counts = Counter(CANONICAL_CATEGORIES.get(str(row[category_key]).strip().casefold(), str(row[category_key]).strip()) for row in rows)
    if sum(counts.values()) != 40 or counts != Counter(EXPECTED):
        raise ValueError(f"Unexpected Golden Dataset distribution: {dict(counts)}")
    return counts
